from mpl_toolkits.mplot3d import Axes3D
import matplotlib.pyplot as plt

import xlrd
from openpyxl import load_workbook

import math

file = xlrd.open_workbook('data_train_PNN.xlsx')
file2 = load_workbook('data_test_PNN.xlsx')
filex = xlrd.open_workbook('data_test_PNN.xlsx')

sheet = file.sheet_by_index(0)
testSheet = file2.get_sheet_by_name('data_test_PNN')
nsheet = filex.sheet_by_index(0)

fig = plt.figure()
fig2 = plt.figure()
ax = fig.add_subplot(111, projection='3d')
ax2 = fig2.add_subplot(111, projection ='3d')

x1 = []
x2 = []
x3 = []
label = []
color = []

data1=list()
data2=list()
data3=list()

temp = []
tempAcc =[]

def train(x,y):
    dt = []
    record = []
    for i in range(x,sheet.nrows-y):
        Xi = sheet.cell(rowx=i,colx=0).value
        x1.append(Xi)
        record.append(Xi)
        Yi = sheet.cell(rowx=i,colx=1).value
        x2.append(Yi)
        record.append(Yi)
        Zi = sheet.cell(rowx=i,colx=2).value
        x3.append(Zi)
        record.append(Zi)
        a = sheet.cell(rowx=i,colx=3).value
        label.append(a)
        record.append(a)
        dt.append(record)
        record = []
    return dt

data1 = train(0,100)
data2 = train(50,50)
data3 = train(100,0)
dataTr = [data1, data2, data3]
length = len(dataTr) -1


def calcTrainMain(smin,smax):
    global dataTr
    global length
    global tempAcc
    global temp
    while (smin<=smax):
        tempAcc = []
        for a in range(length+1):
            dSet = dataTr[a]
            dTrain =[]
            b = 0
            sk = []
            while (b<=length):
                if b == a:
                    pass
                else:
                    dTrain.append(dataTr[b])
                    n = dataTr[b]
                    for r in range(len(dataTr[b])):
                        sk.append(n[r])
                b += 1
            x1 = []
            x2 = []
            x3 = []
            x1tot = 0
            x2tot = 0
            x3tot = 0
            for i in range(len(label)):
                if label[i] == 0:
                    x1tot += 1
                elif label[i] == 1:
                    x2tot += 1
                else:
                    x3tot += 1
            calcGx(x1tot,x2tot,x3tot,smin,sk,dSet)
            accu = acc(dSet)
            tempAcc.append(accu)
        aga = sum(tempAcc)/len(tempAcc)
        temp.append((aga,smin))
        #print('smin',smin,'punya accu',tempAcc)
        for s in (dSet):
            del s[4]
        smin +=0.2
    print(temp)

def calcGx(a,b,c,s,sk,dSet):
    for ii in (dSet):
        if (len(ii) != 4):
            del ii[4]
        f0 = 0
        f1 = 0
        f2 = 0
        Xtest = ii[0]
        Ytest = ii[1]
        Ztest = ii[2]
        for j in range(len(sk)):
            result = math.exp(-(((Xtest - x1[j]) ** 2) + ((Ytest - x2[j]) ** 2) + ((Ztest - x3[j]) ** 2) / 2 * (s ** 2))) #Calc g(x)
            # print(j, ",",result)
            if label[j] == 0:
                f0 += result
            elif label[j] == 1:
                f1 += result
            elif label[j] == 2:
                f2 += result
        #print(f0, " ", f1, " ", f2)
        f0 = f0 / a
        f1 = f1 / b
        f2 = f2 / c
        if f0 > f1 and f0 > f2:
            ii.append(0.0)
        elif f1 > f0 and f1 > f2:
            ii.append(1.0)
        elif f2 > f0 and f2 > f1:
            ii.append(2.0)
        #print('data set:',dSet)



def acc(ts):
    trueValue = 0
    for x in ts:
        if x[3] == x[4]:
            trueValue += 1
    accuracy = float(trueValue) / len(ts) * 100
    return accuracy


def main(acrs):
    dt = []
    x1,x2,x3 = [],[],[]
    record = []
    for i in range(sheet.nrows):
        Xi = sheet.cell(rowx=i,colx=0).value
        x1.append(Xi)
        record.append(Xi)
        Yi = sheet.cell(rowx=i,colx=1).value
        x2.append(Yi)
        record.append(Yi)
        Zi = sheet.cell(rowx=i,colx=2).value
        x3.append(Zi)
        record.append(Zi)
        a = sheet.cell(rowx=i,colx=3).value
        label.append(a)
        record.append(a)
        dt.append(record)
        record = []

    listTable = list()
    for x in range(nsheet.nrows):
        for y in range(nsheet.ncols):
            record.append(nsheet.cell(x, y).value)
        listTable.append(record)
        record =[]

    #######2#######
    x1tot = 0
    x2tot = 0
    x3tot = 0
    for i in range(len(dt)):
        if label[i] == 0:
            x1tot += 1
        elif label[i] == 1:
            x2tot += 1
        else:
            x3tot += 1
    ######3#######
    row = 1
    for ii in (listTable):
        f0 = 0
        f1 = 0
        f2 = 0
        Xtest = ii[0]
        Ytest = ii[1]
        Ztest = ii[2]
        for j in range(149):
            result = math.exp(-(((Xtest - x1[j]) ** 2) + ((Ytest - x2[j]) ** 2) + ((Ztest - x3[j]) ** 2) / 2 * (acrs ** 2))) #Calc g(x)
            # print(j, ",",result)
            if label[j] == 0:
                f0 += result
            elif label[j] == 1:
                f1 += result
            elif label[j] == 2:
                f2 += result
        f0 = f0 / x1tot
        f1 = f1 / x2tot
        f2 = f2 / x3tot
        if f0 > f1 and f0 > f2:
            ii.append(0.0)
            testSheet.cell(row = row,column = 4).value = 0.0
        elif f1 > f0 and f1 > f2:
            ii.append(1.0)
            testSheet.cell(row=row, column=4).value = 1.0
        elif f2 > f0 and f2 > f1:
            ii.append(2.0)
            testSheet.cell(row=row, column=4).value = 2.0
        row +=1
    return listTable


##############MAIN#######################
sett = train(0,0)
rfile = open('Result.txt','w')
x1,x2,x3 =[],[],[]
for x in (sett):
    Xi = x[0]
    x1.append(Xi)
    Yi = x[1]
    x2.append(Yi)
    Zi = x[2]
    x3.append(Zi)
    a = x[3]
    label.append(a)
    if a == 0:
        ax.scatter(Xi, Yi, Zi, s=100, c='red')
    elif a == 1:
        ax.scatter(Xi, Yi, Zi, s=100, c='green')
    else:
        ax.scatter(Xi, Yi, Zi, s=100, c='blue')
    ax.set_xlabel('x1')
    ax.set_ylabel('x2')
    ax.set_zlabel('x3')
#####
calcTrainMain(0,50)
hasil = max(temp)
sigma = hasil[1]
print('hasil akurasi',hasil)
print('sigmanya',sigma)
print(len(x1))
m = main(sigma)
#####
#TXT#
#####
init = 1
for line in m:
    rfile.write(str(init)+'\t')
    for s in line:
        rfile.write(str(s)+'\t')
    rfile.write('\n')
    init +=1
#####
for x in m:
    Xi = x[0]
    x1.append(Xi)
    Yi = x[1]
    x2.append(Yi)
    Zi = x[2]
    x3.append(Zi)
    ase = x[3]
    label.append(ase)
    if ase == 0.0:
        ax2.scatter(Xi, Yi, Zi, s=100, c='red')
    elif ase == 1.0:
        ax2.scatter(Xi, Yi, Zi, s=100, c='green')
    else:
        ax2.scatter(Xi, Yi, Zi, s=100, c='blue')
    ax2.set_xlabel('x1')
    ax2.set_ylabel('x2')
    ax2.set_zlabel('x3')
plt.show()

print(len(main(sigma)))
print('Result',main(sigma))
print(len(x1))
file2.save("Final Result Tugas 1.3 MacLearn Fixed.xlsx")